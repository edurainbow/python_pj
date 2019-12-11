#__author:"Peter"
#date:2019/12/11
from openpyxl import Workbook
from openpyxl import load_workbook

def excel_parser_return_dict(file = 'test.xlsx', sheel_name = 'Sheet1'):
    data = load_workbook(file)#读取xlsx文件
    table = data[sheel_name]#读取sheet数据
    # print('%-22s %-22s %s' % (table['A1'].value, table['B1'].value, table['C1'].value))
    # print(table.rows)
    excel_dict = {}
    row_location = 0
    for row in table.iter_rows():
        if row_location == 0:#跳过第一行！
            row_location +=1
            continue
        else:
            cell_location = 0
            for cell in row:
                if cell_location == 0:#读取第一列的用户名
                    tmp_user = cell.value
                    cell_location += 1
                elif cell_location == 1:#读取第二列的密码
                    tmp_pass = cell.value
                    cell_location += 1
                elif cell_location == 2:#读取第三列的级别
                    tmp_priv = cell.value
                    cell_location += 1
            excel_dict[tmp_user] = tmp_pass,tmp_priv#写入字典
        row_location += 1
    return excel_dict#返回字典
if __name__ == '__main__':
    print(excel_parser_return_dict('test.xlsx'))
